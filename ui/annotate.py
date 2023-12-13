"""
#=====================================================================
# annotate.py
#=====================================================================
# Code that handles annotating a copy of a checklist, using data in
# the Roster
#
# Author: Aidan McNay
# Date: December 8th, 2023
"""

import os
import shutil
from typing import Tuple, Set, List, cast

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from obj.roster_obj import Roster
from obj.roster_entry_obj import RosterEntry, ERROR, WARNING, VALID
from obj.roster_entry_obj import req_types
from obj.coordinates_obj import Coordinates
import exceptions as excp

#---------------------------------------------------------------------
# Define cell fills for each validity
#---------------------------------------------------------------------

error_fill   = PatternFill( patternType = "solid", fgColor = "cc3300" )
warning_fill = PatternFill( patternType = "solid", fgColor = "ffcc00" )
valid_fill   = PatternFill( patternType = "solid", fgColor = "99cc33" )

#---------------------------------------------------------------------
# Wrapper Functions for interacting with an OpenPyXL Worksheet
#---------------------------------------------------------------------

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# Coordinate Manipulation
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# Provides semantic wrappers around OpenPyXL coordinates (ex. 'A5'),
# as well as how they relate to (row, column) tuples

def coord_to_tuple( coordinate: str ) -> Tuple[int, int]:
    """Turns a coordinate ('A5') into a tuple ( 5, 1 )"""
    return openpyxl.utils.cell.coordinate_to_tuple( coordinate )

def tuple_to_coord( idx_tuple: Tuple[ int, int ] ) -> str:
    """Turns a tuple ( 5, 1 ) into a coordinate ('A5')"""
    row = idx_tuple[0]
    col = idx_tuple[1]

    row_str = str( row )
    col_str = openpyxl.utils.cell.get_column_letter( col )

    return col_str + row_str

def cl( coordinate: str ) -> str:
    """Gets the coordinates to the left"""

    row, column = coord_to_tuple( coordinate )
    new_coord_tuple = ( row, column - 1 )
    return tuple_to_coord( new_coord_tuple )

def cr( coordinate: str ) -> str:
    """Gets the coordinates to the left"""

    row, column = coord_to_tuple( coordinate )
    new_coord_tuple = ( row, column + 1 )
    return tuple_to_coord( new_coord_tuple )

def cu( coordinate: str ) -> str:
    """Gets the coordinates to the left"""

    row, column = coord_to_tuple( coordinate )
    new_coord_tuple = ( row - 1, column )
    return tuple_to_coord( new_coord_tuple )

def cd( coordinate: str ) -> str:
    """Gets the coordinates to the left"""

    row, column = coord_to_tuple( coordinate )
    new_coord_tuple = ( row + 1, column )
    return tuple_to_coord( new_coord_tuple )

def coord_to_coordinates( coordinate: str ) -> Coordinates:
    """Turns a coordinate ('A5') into a Coordinates object"""
    row, column = coord_to_tuple( coordinate )

    # Subtract 1 to index from 0
    return Coordinates( row - 1, column - 1 )

#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -
# Cell Manipulation
#- - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - - -

def find_cell( ws: Worksheet, val: str, case_insensitive: bool = True ) -> List[str]:
    """Returns a list of OpenPyXL coordinates (ex. 'A5') for all matching cells"""

    result = []

    for row in ws.rows:
        for cell in row:
            if (                      ( val         in str( cell.value )           ) or
               ( case_insensitive and ( val.lower() in str( cell.value ).lower() ) ) ):
                result.append( cell.coordinate )

    return result

def find_cell_multival( ws: Worksheet, vals: Set[str], case_insensitive: bool = True ) -> List[str]:
    """
    Returns a list of OpenPyXL coordinates (ex. 'A5') for all matching cells with any value in vals
    """

    result = []
    for val in vals:
        result += find_cell( ws, val, case_insensitive )
    return result

def get_val( ws: Worksheet, coordinate: str ) -> str:
    """Returns the value that a cell at a given coordinate holds"""

    return str( ws[ coordinate ].value )

def fill_cell( ws: Worksheet, coordinate: str, fill_pattern: PatternFill ) -> None:
    """Fills a cell with the given fill pattern"""
    ws[ coordinate ].fill = fill_pattern

#---------------------------------------------------------------------
# Roster Interactions
#---------------------------------------------------------------------

def xlsx_copy( roster: Roster, dest_dir: str ) -> str:
    """
    Copies the checklist to the given destination directory, and returns the 
    path to the copied file
    """

    dest_path = os.path.join( dest_dir, f"{roster.netid}.xlsx" )
    shutil.copyfile( roster.filepath, dest_path )
    return dest_path

def color_cell( ws: Worksheet, coord: str, validity_level: int ) -> None:
    """Colors a cell according to its validity"""

    if validity_level == VALID:
        fill_cell( ws, coord, valid_fill )

    if validity_level == WARNING:
        fill_cell( ws, coord, warning_fill )

    if validity_level == ERROR:
        fill_cell( ws, coord, error_fill )

def make_annotated_checklist( roster: Roster, dest_dir: str ) -> None:
    """Makes an annotated checklist in the specified directory"""

    dest_file = xlsx_copy( roster, dest_dir )
    wb = openpyxl.load_workbook( dest_file )
    ws = cast( Worksheet, wb.active )

    # Color the requirements

    req_coords = find_cell_multival( ws, req_types )
    for coord in req_coords:
        req = get_val( ws, coord ).upper()
        entry = cast( RosterEntry, roster.get_entry( coord_to_coordinates( coord ) ) )

        color_cell( ws, coord,                         entry.get_val( "req"    ) ) # Requirement
        color_cell( ws, cr( coord ),                   entry.get_val( "course" ) ) # Course
        color_cell( ws, cr( cr( coord ) ),             entry.get_val( "cred"   ) ) # Credits
        color_cell( ws, cr( cr( cr( coord ) ) ),       entry.get_val( "term"   ) ) # Term
        color_cell( ws, cr( cr( cr( cr( coord ) ) ) ), entry.get_val( "grade"  ) ) # Grade

        if req == "LS": # Also need to color the category
            color_cell( ws, cr( cr( cr( cr( cr( coord ) ) ) ) ), entry.get_val( "cat" ) )

    # Color the checkoffs

    adv_prog_coord_list  = find_cell( ws, "Adv. Programming" )
    tech_writ_coord_list = find_cell( ws, "Tech. Writing"    )

    if len( adv_prog_coord_list ) != 1:
        raise excp.checklist_exceptions.MultipleAttributeError( "Adv. Programming",
                                                                len( adv_prog_coord_list ) )

    if len( tech_writ_coord_list ) != 1:
        raise excp.checklist_exceptions.MultipleAttributeError( "Tech. Writing",
                                                                len( tech_writ_coord_list ) )

    adv_prog_coord  = adv_prog_coord_list [ 0 ]
    tech_writ_coord = tech_writ_coord_list[ 0 ]

    adv_prog_entry  = cast( RosterEntry,
                            roster.get_entry( coord_to_coordinates( adv_prog_coord   ) ) )
    tech_writ_entry = cast( RosterEntry,
                            roster.get_entry( coord_to_coordinates( tech_writ_coord  ) ) )


    color_cell( ws, adv_prog_coord,              adv_prog_entry.get_val( "req"    ) )
    color_cell( ws, cr( cr( adv_prog_coord ) ),  adv_prog_entry.get_val( "course" ) )

    color_cell( ws, tech_writ_coord,             tech_writ_entry.get_val( "req"    ) )
    color_cell( ws, cr( cr( tech_writ_coord ) ), tech_writ_entry.get_val( "course" ) )

    # Save the file
    wb.save( dest_file )
